"""Deterministic, zero-LLM buyer artifacts (register items F1–F4).

The pipeline already COMPUTES everything in here and then throws most of it away: the
six-axis :class:`~prospector.models.ScoreResult`, the Python-computed financial model, and
the :class:`~prospector.models.PriceAnchor` comparables fetched by the `price_comparables`
check. Buyers get eight markdown files and one HTML render; the numbers behind them never
leave the dossier. This module renders them into the machine-readable and re-modellable
formats a buyer can actually work with:

* **F1** ``scorecard.json`` / ``scorecard.csv`` — per-axis score, weight, weighted
  contribution and the composite. ``financial.json`` / ``financial.csv`` — the model's
  verified INPUTS and every OUTPUT Python derives from them. ``comparables.json`` /
  ``comparables.csv`` — the cited price anchors, each row carrying its source URL and the
  literal passage it was lifted from (source-or-die applies to shipped artifacts too, so an
  anchor with no URL is quarantined into ``unsourced``, never silently rendered as evidence).
* **F2** ``scorecard_radar.svg`` — a hand-written six-axis radar. No matplotlib, no
  dependency: polar-to-cartesian arithmetic and a string.
* **F3** ``financial_model.xlsx`` — the same inputs as an editable sheet whose revenue/cost
  cells are LIVE ``=``-formulas pointing at the input cells, so a buyer can change a price
  and watch the model move. Requires ``openpyxl``; absent it, the emitter returns ``None``
  and logs why rather than failing the pack.
* **F4** ``pack.pdf`` — headless-Chrome print of the pack HTML that ``pack_html`` already
  renders. Chrome missing or failing returns ``None``; this NEVER raises into a bundle path.

Two rules shape every function here:

1. **Determinism.** Same dossier in, byte-identical bytes out. No ``now()``, no iteration
   over unordered sets, fixed float precision, fixed axis order (``models.SCORE_AXES``).
   The one exception is the XLSX container: ``openpyxl`` stamps zip member mtimes from the
   wall clock, so the SHEET is deterministic while the file bytes are not — the test asserts
   on cell contents, not on the archive.
2. **Params live in config.** Everything tunable is read from a ``pack_data`` dict on the
   Config object via ``getattr``; the module-level ``DEFAULT_*`` values below are that
   block's documented defaults, not logic literals.
"""
from __future__ import annotations

import csv
import io
import json
import math
import os
import subprocess
import tempfile
import time
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple
from xml.sax.saxutils import escape as _xml_escape

from .models import SCORE_AXES
from .telemetry import logger

try:  # pragma: no cover - exercised by whichever branch the venv is in
    import openpyxl as _openpyxl
    _OPENPYXL_ERROR = ""
except ImportError as exc:  # pragma: no cover
    _openpyxl = None
    _OPENPYXL_ERROR = str(exc)


# --- config block defaults (config.yaml `pack_data:`) -----------------------------------
DEFAULT_ENABLED = False
DEFAULT_FORMATS: Tuple[str, ...] = ("json", "csv", "svg")
DEFAULT_CHROME_PATH = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
DEFAULT_PDF_TIMEOUT_S = 120
# How often to look for the printed PDF. Chrome does not exit after printing (see
# `render_pdf`), so this cadence — not the process — is what ends the wait.
DEFAULT_PDF_POLL_INTERVAL_S = 0.25
# The scoring rubric is 0-5 per axis (score.py: "Score six axes 0-5"). Declared here so the
# radar's radial scale and the scorecard's `max_score` cannot drift from each other.
DEFAULT_AXIS_MAX = 5
# The passage excerpt shipped as a comparable's proof. 600 chars is the same window the moat
# itself shows the model (`s.text[:600]`), so the buyer reads exactly what was ruled on.
DEFAULT_SNIPPET_CHARS = 600
DEFAULT_RADAR: Dict[str, Any] = {
    "size": 560,          # viewBox is size x size
    # Gutter outside the plot for the two-line axis labels. Sized against the longest axis
    # name actually in SCORE_AXES ("Money Provability", ~115px at 13px/600) rendered
    # left-anchored off the 30-degree spoke: at margin 108 it ran 1px past the viewBox.
    "margin": 132,
    "grid": "#c9ced6",
    "spoke": "#aab1bb",
    "ink": "#1b1f24",
    "muted": "#5a6470",
    "accent": "#1f6feb",
    "background": "#ffffff",
    "font": "Helvetica, Arial, sans-serif",
}

# Output filenames. The bundle contract lives in bridge.BUNDLE_FILES; these are the names
# this module emits under, so both sides have exactly one spelling to agree on.
SCORECARD_JSON = "scorecard.json"
SCORECARD_CSV = "scorecard.csv"
FINANCIAL_JSON = "financial.json"
FINANCIAL_CSV = "financial.csv"
COMPARABLES_JSON = "comparables.json"
COMPARABLES_CSV = "comparables.csv"
RADAR_SVG = "scorecard_radar.svg"
FINANCIAL_XLSX = "financial_model.xlsx"
PACK_PDF = "pack.pdf"

# The financial model's INPUT keys, in the order `artifacts._render_financial_model` reads
# them. (key, human label, unit). Ordered tuple, not a dict comprehension over the payload:
# the sheet's row numbers are formula targets, so the order is part of the contract.
FINANCIAL_INPUTS: Tuple[Tuple[str, str, str], ...] = (
    ("monthly_price", "Monthly price", "currency"),
    ("target_customers_month_1", "Customers, month 1", "count"),
    ("target_customers_month_12", "Customers, month 12", "count"),
    ("estimated_cac_gbp", "Customer acquisition cost (CAC)", "currency"),
    ("estimated_clv_gbp", "Customer lifetime value (CLV, stated)", "currency"),
    ("estimated_monthly_churn_pct", "Monthly churn", "percent"),
    ("cost_of_goods_pct", "Cost of goods (COGS)", "percent"),
    ("overhead_month_1_gbp", "Overhead, month 1", "currency"),
    ("sales_cycle_months", "Sales cycle", "months"),
    ("payback_months", "Payback (stated)", "months"),
    ("ltv_cac_ratio", "LTV:CAC (stated)", "ratio"),
    # Appended 2026-08-14, after the tuple's original members, because the sheet's row
    # numbers are formula targets: `financial_xlsx` builds `row_of` by iterating THIS
    # tuple, so appending shifts nothing that already had a row.
    ("repeat_purchases_per_customer", "Repeat sales per buyer", "count"),
)

_FINANCIAL_OUTPUT_LABELS: Dict[str, str] = {
    "revenue_month_1": "Revenue, month 1",
    "revenue_month_12": "Revenue, month 12",
    "growth_multiple_m1_m12": "Growth, M1 to M12",
    "gross_margin_pct": "Gross margin",
    "gross_margin_per_customer": "Gross margin per customer / month",
    "customer_lifetime_value": "Customer lifetime value (derived)",
    "payback_months": "Payback period",
    "ltv_cac_ratio": "LTV:CAC ratio",
    "month_1_cogs": "Month 1 COGS",
    "month_1_gross_profit": "Month 1 gross profit",
    "month_1_overhead": "Month 1 overhead",
    "month_1_net": "Month 1 net",
}


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def settings(cfg: Any = None) -> Dict[str, Any]:
    """Resolve the `pack_data` config block over its defaults.

    Reads through ``getattr`` so an older Config without the block behaves exactly as an
    explicit ``enabled: false`` — this feature must be inert, not fatal, on a stale config.
    """
    raw = getattr(cfg, "pack_data", {}) or {}
    if not isinstance(raw, dict):
        logger.warning("pack_data config is not a mapping; ignoring",
                       extra={"type": type(raw).__name__})
        raw = {}
    formats = raw.get("formats")
    if not isinstance(formats, (list, tuple)) or not formats:
        formats = DEFAULT_FORMATS
    radar = dict(DEFAULT_RADAR)
    if isinstance(raw.get("radar"), dict):
        radar.update(raw["radar"])
    return {
        "enabled": bool(raw.get("enabled", DEFAULT_ENABLED)),
        "formats": tuple(str(f).lower() for f in formats),
        "chrome_path": str(raw.get("chrome_path") or DEFAULT_CHROME_PATH),
        "pdf_timeout_s": int(raw.get("pdf_timeout_s") or DEFAULT_PDF_TIMEOUT_S),
        "pdf_poll_interval_s": float(raw.get("pdf_poll_interval_s")
                                     or DEFAULT_PDF_POLL_INTERVAL_S),
        "axis_max": int(raw.get("axis_max") or DEFAULT_AXIS_MAX),
        "snippet_chars": int(raw.get("snippet_chars") or DEFAULT_SNIPPET_CHARS),
        "radar": radar,
    }


# ---------------------------------------------------------------------------
# Small deterministic helpers
# ---------------------------------------------------------------------------

def _dumps(payload: Any) -> str:
    """JSON with a fixed shape: 2-space indent, insertion order, trailing newline."""
    return json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=False) + "\n"


def _csv(rows: Sequence[Sequence[Any]]) -> str:
    """CSV with an explicit ``\\n`` terminator so the bytes don't depend on the platform."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for row in rows:
        writer.writerow(["" if v is None else v for v in row])
    return buf.getvalue()


def _num(value: Any) -> Optional[float]:
    """Coerce to float, or None. A model that emitted ``"49"`` must not poison the sheet."""
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _round(value: Optional[float], places: int = 4) -> Optional[float]:
    return None if value is None else round(value, places)


def _pretty_axis(axis: str) -> str:
    return axis.replace("_", " ").title()


# ---------------------------------------------------------------------------
# F1a — scorecard
# ---------------------------------------------------------------------------

def scorecard(score: Any, weights: Mapping[str, float],
              *, axis_max: int = DEFAULT_AXIS_MAX) -> Dict[str, Any]:
    """The six-axis ScoreResult as a data structure: score, weight, contribution, composite.

    ``score`` may be ``None`` (a KILL never gets scored, and the artifact path can run before
    scoring). That is reported as ``score_available: false`` with an empty ``axes`` list —
    the honest rendering of "not scored". Inventing zeros here would ship a dossier-shaped
    lie, which is the exact defect `ScoreResult.score_failed` exists to prevent.
    """
    scores = dict(getattr(score, "scores", None) or {}) if score is not None else {}
    justification = dict(getattr(score, "justification", None) or {}) if score is not None else {}
    available = score is not None and bool(scores)

    axes: List[Dict[str, Any]] = []
    contribution_total = 0.0
    for axis in SCORE_AXES:
        if not available:
            continue
        raw = _num(scores.get(axis)) or 0.0
        weight = _num(weights.get(axis)) or 0.0
        contribution = raw * weight
        contribution_total += contribution
        axes.append({
            "axis": axis,
            "label": _pretty_axis(axis),
            "score": _round(raw, 4),
            "max_score": axis_max,
            "weight": _round(weight, 4),
            "weighted_contribution": _round(contribution, 4),
            "justification": str(justification.get(axis, "")),
        })

    stated = _num(getattr(score, "composite", None)) if score is not None else None
    return {
        "score_available": available,
        "score_failed": bool(getattr(score, "score_failed", False)) if score is not None else False,
        "axis_max": axis_max,
        "axes": axes,
        "composite": _round(stated if stated is not None else contribution_total, 4),
        "composite_recomputed": _round(contribution_total, 4),
        "composite_max": _round(
            float(axis_max) * sum(_num(weights.get(a)) or 0.0 for a in SCORE_AXES), 4),
    }


def scorecard_csv(card: Mapping[str, Any]) -> str:
    rows: List[Sequence[Any]] = [
        ["axis", "label", "score", "max_score", "weight", "weighted_contribution"]
    ]
    for row in card.get("axes") or []:
        rows.append([row["axis"], row["label"], row["score"], row["max_score"],
                     row["weight"], row["weighted_contribution"]])
    rows.append(["composite", "Composite", card.get("composite"),
                 card.get("composite_max"), "", card.get("composite_recomputed")])
    return _csv(rows)


# ---------------------------------------------------------------------------
# F1b — financial model
# ---------------------------------------------------------------------------

def financial_model(assumptions: Optional[Mapping[str, Any]]) -> Dict[str, Any]:
    """Inputs plus every figure Python derives from them.

    Mirrors `artifacts._render_financial_model` exactly — same formulas, same precedence
    (a model-STATED payback/CLV/LTV:CAC wins over the derived one, and both are reported so
    a buyer can see when the model contradicted its own arithmetic). Every field is
    None-safe: a business with no declared price renders a partial model with explicit gaps,
    never a wrong total.
    """
    src: Mapping[str, Any] = assumptions or {}
    inputs: Dict[str, Any] = {}
    for key, _label, _unit in FINANCIAL_INPUTS:
        inputs[key] = _round(_num(src.get(key)), 4)

    price = inputs["monthly_price"]
    cust_1 = inputs["target_customers_month_1"]
    cust_12 = inputs["target_customers_month_12"]
    cac = inputs["estimated_cac_gbp"]
    clv_stated = inputs["estimated_clv_gbp"]
    churn = inputs["estimated_monthly_churn_pct"]
    cogs_pct = inputs["cost_of_goods_pct"]
    overhead = inputs["overhead_month_1_gbp"]
    payback_stated = inputs["payback_months"]
    ltv_cac_stated = inputs["ltv_cac_ratio"]

    rev_1 = price * cust_1 if price is not None and cust_1 is not None else None
    rev_12 = price * cust_12 if price is not None and cust_12 is not None else None
    growth = rev_12 / rev_1 if rev_1 not in (None, 0) and rev_12 is not None else None
    gross_margin_pct = 100.0 - cogs_pct if cogs_pct is not None else None
    gm_per_customer = (price * gross_margin_pct / 100.0
                       if price is not None and gross_margin_pct is not None else None)

    # Same gate as the renderer, from the SAME function: `price / churn` is a
    # subscription formula, and applying it to a one-off sale is how a £24 book acquired a
    # £480 lifetime value on 17 of the 68 packs on disk. Imported rather than restated —
    # this module already mirrors the renderer's formulas, and a second copy of the rule is
    # exactly how the sheet and the document come to disagree.
    from .artifacts import revenue_shape
    shape = revenue_shape(src, churn)
    repeat = inputs.get("repeat_purchases_per_customer")
    if shape == "recurring" and price is not None and churn not in (None, 0):
        clv_derived = price / (churn / 100.0)
    elif shape == "one_off" and price is not None and repeat:
        clv_derived = price * repeat
    else:
        clv_derived = None
    clv = clv_stated if clv_stated is not None else clv_derived

    payback_derived = (cac / gm_per_customer
                       if cac is not None and gm_per_customer not in (None, 0) else None)
    payback = payback_stated if payback_stated is not None else payback_derived

    ltv_cac_derived = (clv / cac if clv is not None and cac not in (None, 0) else None)
    ltv_cac = ltv_cac_stated if ltv_cac_stated is not None else ltv_cac_derived

    m1_cogs = rev_1 * cogs_pct / 100.0 if rev_1 is not None and cogs_pct is not None else None
    m1_gross = rev_1 - m1_cogs if rev_1 is not None and m1_cogs is not None else None
    m1_net = m1_gross - overhead if m1_gross is not None and overhead is not None else None

    outputs: Dict[str, Any] = {
        "revenue_month_1": _round(rev_1, 2),
        "revenue_month_12": _round(rev_12, 2),
        "growth_multiple_m1_m12": _round(growth, 4),
        "gross_margin_pct": _round(gross_margin_pct, 4),
        "gross_margin_per_customer": _round(gm_per_customer, 2),
        "customer_lifetime_value": _round(clv, 2),
        "payback_months": _round(payback, 2),
        "ltv_cac_ratio": _round(ltv_cac, 4),
        "month_1_cogs": _round(m1_cogs, 2),
        "month_1_gross_profit": _round(m1_gross, 2),
        "month_1_overhead": _round(overhead, 2),
        "month_1_net": _round(m1_net, 2),
    }

    notes = [str(a) for a in (src.get("assumptions") or []) if str(a).strip()]
    weaknesses = [str(w) for w in (src.get("weaknesses") or []) if str(w).strip()]
    return {
        "model_available": bool(src),
        "revenue_shape": shape,
        "inputs": inputs,
        "input_labels": {k: label for k, label, _u in FINANCIAL_INPUTS},
        "outputs": outputs,
        "derived": {
            "customer_lifetime_value": _round(clv_derived, 2),
            "payback_months": _round(payback_derived, 2),
            "ltv_cac_ratio": _round(ltv_cac_derived, 4),
        },
        "assumptions": notes,
        "weaknesses": weaknesses,
    }


def financial_csv(model: Mapping[str, Any]) -> str:
    labels = model.get("input_labels") or {}
    rows: List[Sequence[Any]] = [["section", "key", "label", "value"]]
    for key, _label, _unit in FINANCIAL_INPUTS:
        rows.append(["input", key, labels.get(key, key), (model.get("inputs") or {}).get(key)])
    for key, value in (model.get("outputs") or {}).items():
        rows.append(["output", key, _FINANCIAL_OUTPUT_LABELS.get(key, key), value])
    for note in model.get("assumptions") or []:
        rows.append(["assumption", "", "", note])
    for note in model.get("weaknesses") or []:
        rows.append(["weakness", "", "", note])
    return _csv(rows)


# ---------------------------------------------------------------------------
# F1c — price comparables
# ---------------------------------------------------------------------------

def comparables(anchors: Sequence[Any], sources: Sequence[Any],
                *, snippet_chars: int = DEFAULT_SNIPPET_CHARS) -> Dict[str, Any]:
    """Price anchors with their URL and the literal passage each was lifted from.

    Source-or-die applies to a shipped artifact exactly as it does to a verdict: an anchor
    with no URL and no retrievable passage is not evidence, so it goes to ``unsourced`` with
    a reason instead of being rendered as a comparable a buyer could quote. It is kept rather
    than dropped — the receipt that the rail ran is the point.
    """
    by_id: Dict[str, Any] = {}
    for s in sources or []:
        sid = str(getattr(s, "source_id", "") or "")
        if sid and sid not in by_id:
            by_id[sid] = s

    rows: List[Dict[str, Any]] = []
    unsourced: List[Dict[str, Any]] = []
    for anchor in anchors or []:
        source_id = str(getattr(anchor, "source_id", "") or "")
        source = by_id.get(source_id)
        url = str(getattr(anchor, "url", "") or "") or str(getattr(source, "url", "") or "")
        text = str(getattr(source, "text", "") or "")
        snippet = text[:snippet_chars]
        row = {
            "amount": _round(_num(getattr(anchor, "amount", None)), 4),
            "currency": str(getattr(anchor, "currency", "") or ""),
            "cadence": str(getattr(anchor, "cadence", "") or "unknown"),
            "what": str(getattr(anchor, "what", "") or ""),
            "amount_pence_gbp": getattr(anchor, "amount_pence_gbp", None),
            "source_id": source_id,
            "url": url,
            "cited_snippet": snippet,
        }
        if not url or not snippet:
            reason = ("no source URL on the anchor" if not url
                      else "the cited passage is not in this dossier")
            unsourced.append({**row, "reason": reason})
            continue
        rows.append(row)

    return {
        "count": len(rows),
        "anchors": rows,
        "unsourced": unsourced,
        "note": ("Every row carries the URL it came from and the literal passage the price "
                 "was read out of. Anchors we could not re-cite are listed under "
                 "'unsourced' rather than shown as evidence."),
    }


def comparables_csv(payload: Mapping[str, Any]) -> str:
    rows: List[Sequence[Any]] = [[
        "amount", "currency", "cadence", "what", "amount_pence_gbp",
        "source_id", "url", "cited_snippet",
    ]]
    for row in payload.get("anchors") or []:
        rows.append([row["amount"], row["currency"], row["cadence"], row["what"],
                     row["amount_pence_gbp"], row["source_id"], row["url"],
                     row["cited_snippet"]])
    return _csv(rows)


# ---------------------------------------------------------------------------
# F2 — radar SVG, hand-written
# ---------------------------------------------------------------------------

def radar_svg(card: Mapping[str, Any], *, title: str = "",
              radar: Optional[Mapping[str, Any]] = None) -> str:
    """A six-axis radar as inline SVG. No plotting library, no dependency, no raster.

    Polar-to-cartesian only: axis *i* of *n* sits at ``-90° + i·360/n`` so the first axis is
    at twelve o'clock, and its vertex radius is ``(score / axis_max) · R``. Rendered for
    white paper (the pack prints), with an explicit white background rect so a dark-mode
    viewer does not serve black-on-black.
    """
    conf = dict(DEFAULT_RADAR)
    if radar:
        conf.update(radar)
    size = float(conf["size"])
    margin = float(conf["margin"])
    axis_max = float(card.get("axis_max") or DEFAULT_AXIS_MAX) or 1.0
    axes = list(card.get("axes") or [])
    cx = cy = size / 2.0
    radius = max(size / 2.0 - margin, 1.0)
    n = len(axes)

    def point(index: int, r: float) -> Tuple[float, float]:
        angle = -math.pi / 2.0 + (2.0 * math.pi * index / n)
        return cx + r * math.cos(angle), cy + r * math.sin(angle)

    def fmt(x: float, y: float) -> str:
        return f"{x:.2f},{y:.2f}"

    parts: List[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {size:.0f} {size:.0f}" '
        f'width="{size:.0f}" height="{size:.0f}" role="img" '
        f'aria-label="Six-axis opportunity scorecard">',
        f'<rect x="0" y="0" width="{size:.0f}" height="{size:.0f}" '
        f'fill="{conf["background"]}"/>',
    ]
    if title:
        parts.append(
            f'<text x="{cx:.2f}" y="28" text-anchor="middle" font-family="{conf["font"]}" '
            f'font-size="17" font-weight="600" fill="{conf["ink"]}">{_xml_escape(title)}</text>'
        )

    if not axes:
        parts.append(
            f'<text x="{cx:.2f}" y="{cy:.2f}" text-anchor="middle" '
            f'font-family="{conf["font"]}" font-size="15" fill="{conf["muted"]}">'
            'Not scored — no ScoreResult on this dossier</text>')
        parts.append("</svg>")
        return "\n".join(parts) + "\n"

    # Grid rings, one per rubric point, plus the ring labels up the twelve o'clock spoke.
    for level in range(1, int(axis_max) + 1):
        r = radius * level / axis_max
        ring = " ".join(fmt(*point(i, r)) for i in range(n))
        parts.append(f'<polygon points="{ring}" fill="none" stroke="{conf["grid"]}" '
                     f'stroke-width="1"/>')
        parts.append(
            f'<text x="{cx + 5:.2f}" y="{cy - r + 4:.2f}" font-family="{conf["font"]}" '
            f'font-size="10" fill="{conf["muted"]}">{level}</text>')

    # Spokes.
    for i in range(n):
        ex, ey = point(i, radius)
        parts.append(f'<line x1="{cx:.2f}" y1="{cy:.2f}" x2="{ex:.2f}" y2="{ey:.2f}" '
                     f'stroke="{conf["spoke"]}" stroke-width="1"/>')

    # The score polygon.
    vertices = [point(i, radius * float(row.get("score") or 0.0) / axis_max)
                for i, row in enumerate(axes)]
    parts.append(
        f'<polygon points="{" ".join(fmt(x, y) for x, y in vertices)}" '
        f'fill="{conf["accent"]}" fill-opacity="0.22" stroke="{conf["accent"]}" '
        f'stroke-width="2" stroke-linejoin="round"/>')
    for x, y in vertices:
        parts.append(f'<circle cx="{x:.2f}" cy="{y:.2f}" r="3.5" fill="{conf["accent"]}"/>')

    # Axis labels + the score values, outside the plot.
    for i, row in enumerate(axes):
        lx, ly = point(i, radius + 20.0)
        angle = -math.pi / 2.0 + (2.0 * math.pi * i / n)
        cos_a = math.cos(angle)
        anchor = "middle" if abs(cos_a) < 0.2 else ("start" if cos_a > 0 else "end")
        label = _xml_escape(str(row.get("label") or row.get("axis") or ""))
        value = f'{row.get("score")}/{row.get("max_score")}'
        weight = row.get("weight")
        parts.append(
            f'<text x="{lx:.2f}" y="{ly:.2f}" text-anchor="{anchor}" '
            f'font-family="{conf["font"]}" font-size="13" font-weight="600" '
            f'fill="{conf["ink"]}">{label}</text>')
        parts.append(
            f'<text x="{lx:.2f}" y="{ly + 15:.2f}" text-anchor="{anchor}" '
            f'font-family="{conf["font"]}" font-size="12" fill="{conf["muted"]}">'
            f'{_xml_escape(value)} &#183; weight {weight}</text>')

    composite = card.get("composite")
    parts.append(
        f'<text x="{cx:.2f}" y="{size - 18:.2f}" text-anchor="middle" '
        f'font-family="{conf["font"]}" font-size="13" fill="{conf["ink"]}">'
        f'Composite {composite} of {card.get("composite_max")}</text>')
    parts.append("</svg>")
    return "\n".join(parts) + "\n"


# ---------------------------------------------------------------------------
# F3 — XLSX with live formulas
# ---------------------------------------------------------------------------

def financial_xlsx(model: Mapping[str, Any], out_path: str,
                   *, title: str = "") -> Optional[str]:
    """Write the financial model as an editable sheet whose outputs are LIVE formulas.

    The point of the format is re-modelling: every output cell is an ``=``-expression over
    the input cells, so changing the price in B4 moves revenue, margin, payback and LTV:CAC.
    A sheet of baked values would just be the CSV with extra steps.

    Returns the path, or ``None`` with a logged reason — a missing ``openpyxl`` or an
    unwritable path must never take down artifact generation.
    """
    if _openpyxl is None:
        logger.warning("pack_data: xlsx skipped, openpyxl not importable",
                       extra={"error": _OPENPYXL_ERROR})
        return None
    try:
        wb = _openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Model"
        labels = model.get("input_labels") or {}
        inputs = model.get("inputs") or {}

        ws["A1"] = title or "Financial model"
        ws["A2"] = ("Inputs are the verified figures from the dossier. Every output below is "
                    "a live formula over them — change an input and the model re-computes.")
        ws["A4"] = "INPUTS"
        ws["A5"], ws["B5"], ws["C5"] = "Label", "Value", "Key"

        row_of: Dict[str, int] = {}
        r = 6
        for key, _label, _unit in FINANCIAL_INPUTS:
            ws.cell(row=r, column=1, value=labels.get(key, key))
            ws.cell(row=r, column=2, value=inputs.get(key))
            ws.cell(row=r, column=3, value=key)
            row_of[key] = r
            r += 1

        def b(key: str) -> str:
            return f"B{row_of[key]}"

        out_start = r + 2
        ws.cell(row=out_start - 1, column=1, value="OUTPUTS (live formulas)")
        ws.cell(row=out_start, column=1, value="Label")
        ws.cell(row=out_start, column=2, value="Value")
        ws.cell(row=out_start, column=3, value="Formula shown")

        price, c1, c12 = b("monthly_price"), b("target_customers_month_1"), b("target_customers_month_12")
        cac, churn = b("estimated_cac_gbp"), b("estimated_monthly_churn_pct")
        repeat_cell = b("repeat_purchases_per_customer")
        cogs, overhead = b("cost_of_goods_pct"), b("overhead_month_1_gbp")

        r = out_start + 1
        rev1_cell, rev12_cell = f"B{r}", f"B{r + 1}"
        gm_pct_cell = f"B{r + 3}"
        gm_cust_cell = f"B{r + 4}"
        clv_cell = f"B{r + 5}"
        cogs_cell = f"B{r + 8}"
        gross_cell = f"B{r + 9}"

        formulas: List[Tuple[str, str]] = [
            ("revenue_month_1", f"={price}*{c1}"),
            ("revenue_month_12", f"={price}*{c12}"),
            ("growth_multiple_m1_m12", f'=IF({rev1_cell}=0,"",{rev12_cell}/{rev1_cell})'),
            ("gross_margin_pct", f"=100-{cogs}"),
            ("gross_margin_per_customer", f"={price}*{gm_pct_cell}/100"),
            # One-off businesses have no churn to divide by; their lifetime value is the
            # price times how often the buyer comes back. A single formula for both shapes
            # is the spreadsheet half of the defect fixed in `_render_financial_model`.
            ("customer_lifetime_value",
             (f'=IF({repeat_cell}=0,"",{price}*{repeat_cell})'
              if model.get("revenue_shape") == "one_off"
              else f'=IF({churn}=0,"",{price}/({churn}/100))')),
            ("payback_months", f'=IF({gm_cust_cell}=0,"",{cac}/{gm_cust_cell})'),
            ("ltv_cac_ratio", f'=IF({cac}=0,"",{clv_cell}/{cac})'),
            ("month_1_cogs", f"={rev1_cell}*{cogs}/100"),
            ("month_1_gross_profit", f"={rev1_cell}-{cogs_cell}"),
            ("month_1_overhead", f"={overhead}"),
            ("month_1_net", f"={gross_cell}-{overhead}"),
        ]
        for key, formula in formulas:
            ws.cell(row=r, column=1, value=_FINANCIAL_OUTPUT_LABELS.get(key, key))
            ws.cell(row=r, column=2, value=formula)
            ws.cell(row=r, column=3, value=formula)
            r += 1

        r += 1
        for heading, items in (("Key assumptions", model.get("assumptions") or []),
                               ("Model weaknesses", model.get("weaknesses") or [])):
            if not items:
                continue
            ws.cell(row=r, column=1, value=heading)
            r += 1
            for item in items:
                ws.cell(row=r, column=1, value=item)
                r += 1
            r += 1

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 34
        wb.save(out_path)
        return out_path
    except Exception as exc:  # never break the bundle over a spreadsheet
        logger.warning("pack_data: xlsx generation failed",
                       extra={"error": str(exc), "path": out_path})
        return None


# ---------------------------------------------------------------------------
# F4 — PDF via headless Chrome
# ---------------------------------------------------------------------------

def render_pdf(html: str, out_path: str, *,
               chrome_path: str = DEFAULT_CHROME_PATH,
               timeout_s: int = DEFAULT_PDF_TIMEOUT_S,
               poll_interval_s: float = DEFAULT_PDF_POLL_INTERVAL_S) -> Optional[str]:
    """Print an HTML string to PDF with headless Chrome. Returns the path, or ``None``.

    **We wait for the FILE, not for the process.** Measured on Chrome 151.0.7922.72 /
    macOS 14.5: ``--headless --print-to-pdf`` logs "8205 bytes written to file …", produces a
    valid PDF, and then does not exit — ``timeout 45 … ; rc=124`` with the finished PDF
    already on disk, in all of ``--headless``, ``--headless=old`` and ``--headless=new``.
    A ``subprocess.run(timeout=…)`` therefore burns the whole timeout and then reports
    failure over a PDF that rendered fine, which is exactly what the first version of this
    function did. So: poll for the output, accept it once its size has settled and it starts
    with ``%PDF-``, and kill Chrome ourselves.

    Chrome absent, Chrome failing, Chrome hanging with nothing written, or Chrome exiting 0
    while writing nothing are all the SAME outcome here: a logged reason and ``None``. This
    runs on the publish path, so an exception escaping would turn "no PDF" into "no pack".
    """
    if not html or not html.strip():
        logger.warning("pack_data: pdf skipped, no HTML to print")
        return None
    if not os.path.exists(chrome_path):
        logger.warning("pack_data: pdf skipped, Chrome not found",
                       extra={"chrome_path": chrome_path})
        return None
    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp(prefix="prospector-pdf-")
        src = os.path.join(tmp_dir, "pack.html")
        with open(src, "w", encoding="utf-8") as fh:
            fh.write(html)
        if os.path.exists(out_path):
            os.remove(out_path)   # a stale file must not be mistaken for this render
        cmd = [
            chrome_path,
            "--headless",
            "--disable-gpu",
            "--no-pdf-header-footer",
            "--no-first-run",
            # Its own throwaway profile: printing must never contend with (or mutate) the
            # operator's live Chrome profile, which is also how this stays runnable headless
            # on a machine where Chrome is already open.
            f"--user-data-dir={os.path.join(tmp_dir, 'profile')}",
            f"--print-to-pdf={out_path}",
            f"file://{src}",
        ]
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        deadline = time.monotonic() + timeout_s
        last_size = -1
        settled = 0
        exited: Optional[int] = None
        while time.monotonic() < deadline:
            size = os.path.getsize(out_path) if os.path.exists(out_path) else 0
            settled = settled + 1 if size > 0 and size == last_size else 0
            last_size = size
            if settled >= 2:
                break
            exited = proc.poll()
            if exited is not None:
                break
            time.sleep(poll_interval_s)

        if proc.poll() is None:
            proc.kill()
            proc.wait()
        if exited not in (None, 0) and last_size == 0:
            logger.warning("pack_data: chrome exited non-zero printing the pack",
                           extra={"returncode": exited})
            return None
        if not os.path.exists(out_path) or os.path.getsize(out_path) == 0:
            logger.warning("pack_data: chrome wrote no pdf",
                           extra={"path": out_path, "returncode": exited})
            return None
        with open(out_path, "rb") as fh:
            if fh.read(5) != b"%PDF-":
                logger.warning("pack_data: chrome wrote a file that is not a pdf",
                               extra={"path": out_path})
                return None
        return out_path
    except Exception as exc:
        logger.warning("pack_data: pdf generation failed",
                       extra={"error": str(exc), "path": out_path})
        return None
    finally:
        if tmp_dir:
            try:
                import shutil
                shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception:  # pragma: no cover - cleanup is best-effort
                pass


# ---------------------------------------------------------------------------
# Assembly
# ---------------------------------------------------------------------------

def _anchors_for(candidate: Any) -> List[Any]:
    """Rehydrate the price anchors verify() stashed on the candidate's tags."""
    try:
        from .price_comparables import anchors_from_tags
        return list(anchors_from_tags(candidate))
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("pack_data: could not read price comparables",
                       extra={"error": str(exc)})
        return []


def _sources_for(dossier: Any) -> List[Any]:
    try:
        return list(getattr(dossier, "all_sources", None) or [])
    except Exception:  # pragma: no cover - defensive
        return []


def build_text_artifacts(dossier: Any, cfg: Any = None, *,
                         financial_assumptions: Optional[Mapping[str, Any]] = None,
                         ) -> Dict[str, str]:
    """The json/csv/svg artifacts for one dossier, keyed by filename.

    Honours ``pack_data.formats`` but NOT ``pack_data.enabled`` — the caller owns the
    on/off decision so this stays directly testable and directly usable from a backfill.
    """
    conf = settings(cfg)
    formats = conf["formats"]
    candidate = getattr(dossier, "candidate", None)
    weights = getattr(cfg, "weights", None) or {}

    card = scorecard(getattr(dossier, "score", None), weights, axis_max=conf["axis_max"])
    if financial_assumptions is None:
        tags = (getattr(candidate, "tags", None) or {})
        raw = tags.get("financial_assumptions")
        financial_assumptions = raw if isinstance(raw, dict) else None
    model = financial_model(financial_assumptions)
    comps = comparables(_anchors_for(candidate), _sources_for(dossier),
                        snippet_chars=conf["snippet_chars"])

    out: Dict[str, str] = {}
    if "json" in formats:
        out[SCORECARD_JSON] = _dumps(card)
        out[FINANCIAL_JSON] = _dumps(model)
        out[COMPARABLES_JSON] = _dumps(comps)
    if "csv" in formats:
        out[SCORECARD_CSV] = scorecard_csv(card)
        out[FINANCIAL_CSV] = financial_csv(model)
        out[COMPARABLES_CSV] = comparables_csv(comps)
    if "svg" in formats:
        out[RADAR_SVG] = radar_svg(card, title=str(getattr(candidate, "title", "") or ""),
                                   radar=conf["radar"])
    return out


def write_artifacts(dossier: Any, cfg: Any, out_dir: str, *,
                    financial_assumptions: Optional[Mapping[str, Any]] = None,
                    pack_html: str = "") -> Dict[str, str]:
    """Write every enabled artifact into ``out_dir``; returns ``{filename: path}``.

    The binary formats (``xlsx``, ``pdf``) only appear when their dependency is present and
    their generator succeeded. A missing entry is the signal; nothing raises.
    """
    conf = settings(cfg)
    os.makedirs(out_dir, exist_ok=True)
    written: Dict[str, str] = {}

    for name, content in build_text_artifacts(
            dossier, cfg, financial_assumptions=financial_assumptions).items():
        path = os.path.join(out_dir, name)
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(content)
        written[name] = path

    if "xlsx" in conf["formats"]:
        if financial_assumptions is None:
            tags = (getattr(getattr(dossier, "candidate", None), "tags", None) or {})
            raw = tags.get("financial_assumptions")
            financial_assumptions = raw if isinstance(raw, dict) else None
        path = financial_xlsx(
            financial_model(financial_assumptions),
            os.path.join(out_dir, FINANCIAL_XLSX),
            title=str(getattr(getattr(dossier, "candidate", None), "title", "") or ""),
        )
        if path:
            written[FINANCIAL_XLSX] = path

    if "pdf" in conf["formats"]:
        path = render_pdf(pack_html, os.path.join(out_dir, PACK_PDF),
                          chrome_path=conf["chrome_path"],
                          timeout_s=conf["pdf_timeout_s"],
                          poll_interval_s=conf["pdf_poll_interval_s"])
        if path:
            written[PACK_PDF] = path

    return written


def artifacts_for_bundle(dossier: Any, cfg: Any = None, *,
                         financial_assumptions: Optional[Mapping[str, Any]] = None,
                         ) -> Dict[str, str]:
    """The bundle wire-in: ``{}`` unless ``pack_data.enabled`` is on.

    Returns text artifacts only, so the caller's ``Dict[str, str]`` artifact map stays a map
    of strings. Binary formats go through :func:`write_artifacts`, which needs a directory.
    """
    conf = settings(cfg)
    if not conf["enabled"]:
        return {}
    return build_text_artifacts(dossier, cfg, financial_assumptions=financial_assumptions)
